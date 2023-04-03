# Parsing by key words, such as aged, solutioned, .......
# Present the extracted in another format

import json

with open(r"..\result.json","r",encoding="utf-8") as f:
    results = json.load(f)

import openpyxl
xls = openpyxl.Workbook()
sht = xls.create_sheet("a",index=0)

num = 0
num_unit = 0
col_i = 1
all_info = dict()
for doi,parags in results.items():
    for source,source_info in parags.items():
        if source=="from_text" and source_info:
            for parags_ in source_info:
                for parag,info in parags_.items():
                    for act_attr in info:
                        for sent,s_info in act_attr.items():
                            for act_info in s_info:
                                if "action" in act_info.keys():
                                    if "aged" in act_info["action"] or "aging" in act_info["action"] or "age" in act_info["action"] or "solutioned" in act_info["action"] or "solutioning" in act_info["action"] or "heat" in act_info["action"] or "treatment" in act_info["action"]:
                                        if not act_info["action"].endswith("temperature"):
                                            num += 1
                                            if "units" in act_info.keys():
                                                if act_info["units"]:
                                                    sht.cell(col_i, 1, str(doi))  # 写入表格的只能是字符型数据
                                                    sht.cell(col_i, 2, str(act_info["action"]))  # 写入表格的只能是字符型数据
                                                    sht.cell(col_i, 3, str(act_info["units"]))  # 写入表格的只能是字符型数据
                                                    if "composition" in act_info.keys():
                                                        sht.cell(col_i, 5, str(act_info["composition"]))
                                                    col_i += 1
                                                    num_unit += 1
        elif source=="from_table" and source_info:
            for u_info in source_info:
                sht.cell(col_i, 1, str(doi))  # 写入表格的只能是字符型数据
                if "age" in u_info.keys():
                    sht.cell(col_i, 2, str("age"))  # 写入表格的只能是字符型数据
                    sht.cell(col_i, 3, str(u_info["age"]))
                elif "solution" in u_info.keys():
                    sht.cell(col_i, 2, str("solution"))  # 写入表格的只能是字符型数据
                    sht.cell(col_i, 3, str(u_info["solution"]))
                sht.cell(col_i, 4, str(u_info["unit"]))  # 写入表格的只能是字符型数据
                if "material" in u_info.keys():
                    sht.cell(col_i, 5, str(u_info["material"]))
                if "other_info" in u_info.keys():
                    sht.cell(col_i, 6, str(u_info["other_info"]))
                col_i += 1
xls.save(r'...\results.xlsx')


