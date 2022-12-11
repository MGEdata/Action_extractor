from sklearn.feature_extraction.text import TfidfVectorizer
import nltk
import re
import pandas as pd
from sklearn.model_selection import train_test_split
from gensim.models.doc2vec import TaggedDocument
import multiprocessing
from sklearn.linear_model import LogisticRegression
import numpy as np
import joblib


def tokenize_text(text):
    tokens = []
    for sent in nltk.sent_tokenize(text):
        for word in nltk.word_tokenize(sent):
            tokens.append(word)
    return tokens


def vec_for_learning(model, tagged_docs):
    sents = tagged_docs.values
    targets, regressors = zip(*[(doc.tags[0], model.infer_vector(doc.words, steps=20)) for doc in sents])
    return targets, regressors


def tfidf_extractor(corpus, ngram_range=(1, 1)):
    vectorizer = TfidfVectorizer(min_df=1, norm="l2", smooth_idf=True, use_idf=True, ngram_range=ngram_range)
    features = vectorizer.fit_transform(corpus)

    return vectorizer, features


def get_vectors(bow_vectorizer, tagged_docs, doc_label):
    sents = tagged_docs.values
    outcome = list()
    # infer_vec = model.infer_vector(doc.words, steps=5)
    for doc in sents:
        doc_list = list()
        doc_text = " ".join(doc.words)
        doc_list.append(doc_text)
        model_infer = bow_vectorizer.transform(doc_list)
        model_infer = model_infer.toarray()[0]

        # if "experiment" in doc_label[doc_text].lower() or "method" in doc_label[doc_text].lower():
        #     tag_vec = np.ones(1)
        #     model_i = np.concatenate((model_infer, tag_vec), axis=0)
        # else:
        #     tag_vec = np.zeros(1)
        #     model_i = np.concatenate((model_infer, tag_vec), axis=0)

        outcome.append((doc.tags[0], model_infer))
    targets, regressors = zip(*outcome)
    return targets, regressors


def new_model_train(corpus_path, save_path, size):
    """
    在确定好模型类别之后，将所有的标记语料用于训练模型，保存模型。
    :return:
    """
    csv_data = pd.read_csv(corpus_path)
    df = pd.DataFrame(csv_data)
    df_shape = df.shape
    outcome = {}
    for i in range(size):
        example_i = df.iloc[i]
        if len(example_i) > 0:
            outcome.update({example_i[0]: example_i[1]})
    paragraph_list = list(outcome.keys())
    words_list = []
    outcome_1 = {}
    all_corpus = list()
    doc_label = dict()
    for i in range(len(paragraph_list)):
        r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        paragraph_0 = re.sub(r, ' ', paragraph_list[i].lower())
        words_list_0 = nltk.word_tokenize(paragraph_0)
        words_list.append(words_list_0)
        keys = " ".join(words_list[i])
        all_corpus.append(keys)
        outcome_1.update({keys: df.iloc[i][1]})
        doc_label.update({keys: df.iloc[i][2]})
    df = pd.DataFrame(pd.Series(outcome_1), columns=['y'])
    train = df.reset_index().rename(columns={'index': 'x'})
    train_tagged = train.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)
    bow_vectorizer, bow_features = tfidf_extractor(all_corpus, ngram_range=(1, 3))
    y_train, X_train = get_vectors(bow_vectorizer, train_tagged, doc_label)
    log = LogisticRegression(max_iter=1000, class_weight='balanced').fit(X_train, y_train)
    joblib.dump(log, save_path)


def get_accuracy_f1(lable_path, random_index, seed_size):
    csv_data = pd.read_csv(lable_path)
    df = pd.DataFrame(csv_data)
    df_shape = df.shape
    length = df_shape[0]
    outcome = {}
    for i in range(seed_size):  # length
        example_i = df.iloc[i]
        if len(example_i) > 0:
            outcome.update({example_i[0]: example_i[1]})
    paragraph_list = list(outcome.keys())
    english_punctuations = [',', '.', ':', ';', '?', '(', ')', '[', ']', '&', '!', '*', '@', '#', '$', '%']
    words_list = []
    outcome_1 = {}
    all_corpus = list()
    doc_label = dict()  # 用于记录每一个自然段对应的标签
    for i in range(len(paragraph_list)):
        r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        paragraph_0 = re.sub(r, ' ', paragraph_list[i].lower())
        words_list_0 = nltk.word_tokenize(paragraph_0)
        words_list.append(words_list_0)
        keys = " ".join(words_list[i])
        all_corpus.append(keys)
        outcome_1.update({keys: df.iloc[i][1]})
        doc_label.update({keys: df.iloc[i][2]})
    # 可以去重
    df = pd.DataFrame(pd.Series(outcome_1), columns=['y'])  # 将字典转化为dataframe的形式,第一列会变成索引列
    df = df.reset_index().rename(columns={'index': 'x'})

    train, test = train_test_split(df, test_size=0.2, random_state=random_index)  # 分为训练集和测试集

    train_tagged = train.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)
    test_tagged = test.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)
    cores = multiprocessing.cpu_count()
    bow_vectorizer, bow_features = tfidf_extractor(all_corpus, ngram_range=(1, 3))

    y_train, X_train = get_vectors(bow_vectorizer, train_tagged, doc_label)


    from sklearn.model_selection import cross_val_score
    log = LogisticRegression(max_iter=1000, class_weight='balanced')
    all_tagged = df.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)
    y_train, X_train = get_vectors(bow_vectorizer, all_tagged, doc_label)
    scores = cross_val_score(log, X_train, y_train, cv=5)
    print(scores)
    ave_score = np.sum(scores) / 5
    print(ave_score)


def predict_on_file(model_path, label_path, target_excel, predict_results_path):
    """
    训练段落分类模型并在新语料中预测，结果输出到新的excel文件中
    :param model_path: 训练好的模型被保存到的地址
    :param label_path: 标记的数据所储存的地址
    :param target_excel: 要预测的新语料所在地址
    :param predict_results_path: 预测结果的输出地址
    :return:
    """
    csv_data = pd.read_csv(label_path)
    df = pd.DataFrame(csv_data)
    df_shape = df.shape
    length = df_shape[0]
    outcome = {}
    for i in range(300):
        example_i = df.iloc[i]
        if len(example_i) > 0:
            outcome.update({example_i[0]: example_i[1]})
    paragraph_list = list(outcome.keys())
    words_list = []
    outcome_1 = {}
    all_corpus = list()
    doc_label = dict()
    for i in range(len(paragraph_list)):
        r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        paragraph_0 = re.sub(r, ' ', paragraph_list[i].lower())
        words_list_0 = nltk.word_tokenize(paragraph_0)
        words_list.append(words_list_0)
        keys = " ".join(words_list[i])
        all_corpus.append(keys)
        outcome_1.update({keys: df.iloc[i][1]})
        doc_label.update({keys: df.iloc[i][2]})

    df = pd.DataFrame(pd.Series(outcome_1), columns=['y'])
    df = df.reset_index().rename(columns={'index': 'x'})
    train = df

    train_tagged = train.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)
    bow_vectorizer, bow_features = tfidf_extractor(all_corpus, ngram_range=(1, 1))
    y_test, X_test = get_vectors(bow_vectorizer, train_tagged, doc_label)
    log = LogisticRegression(max_iter=1000, class_weight='balanced').fit(X_test, y_test)
    joblib.dump(log, model_path)

    # 利用本地向量进行预测
    import xlrd
    new_data = xlrd.open_workbook(target_excel)
    sht = new_data.sheet_by_index(0)
    rows = sht.nrows
    col_0 = sht.col_values(0)
    col_1 = sht.col_values(1)
    outcome = dict()
    outcome_v = list()
    doi_outcome = dict()
    for i in range(rows):
        words_list = []
        example_i = sht.row_values(i)
        if len(example_i) > 0:
            outcome.update({example_i[2]: example_i[1]})
            text = example_i[2]
            r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
            paragraph_ = re.sub(r, ' ', text.lower())
            words_list_0 = nltk.word_tokenize(paragraph_)
            keys = " ".join(words_list_0)
            doi_outcome.update({keys: example_i[0]})
    paragraph_list = list(outcome.keys())
    words_list = list()
    doc_label = dict()

    for i in range(len(paragraph_list)):
        r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        paragraph_0 = re.sub(r, ' ', paragraph_list[i].lower())
        words_list_0 = nltk.word_tokenize(paragraph_0)
        words_list.append(words_list_0)
        keys = " ".join(words_list[i])
        doc_label.update({keys: col_1[i]})

    import openpyxl
    xls = openpyxl.Workbook()
    sht = xls.create_sheet(index=0)
    col_i = 1
    for words in words_list:
        doc_text = " ".join(words)
        doc_list = list()
        doc_list.append(doc_text)
        model_infer = bow_vectorizer.transform(doc_list)
        model_infer = model_infer.toarray()[0]

        # if "experiment" in doc_label[doc_text].lower() or "method" in doc_label[doc_text].lower():
        #     tag_vec = np.ones(1)
        #     model_i = np.concatenate((model_infer, tag_vec), axis=0)
        # else:
        #     tag_vec = np.zeros(1)
        #     model_i = np.concatenate((model_infer, tag_vec), axis=0)
        outcome_v.append(model_infer)
        sht.cell(col_i, 1, paragraph_list[col_i - 1])
        doi = doi_outcome[doc_text]
        sht.cell(col_i, 3, doi)
        col_i += 1
    predict_outcome = log.predict(tuple(outcome_v))
    pre_prob = log.predict_proba(tuple(outcome_v)).tolist()
    col_i = 1
    for tag in predict_outcome:
        sht.cell(col_i, 2, tag)
        sht.cell(col_i, 4, str(pre_prob[col_i - 1]))
        col_i += 1
    xls.save(predict_results_path)

