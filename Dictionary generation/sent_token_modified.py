# -*- coding: utf-8 -*-
"""
Created on Mon Nov 15 16:23:59 2021

@author: win
"""

import nltk
import re
import openpyxl
from dictionary import Dictionary

error_chunks = Dictionary(r"dictionary.ini").error_chunks

# error_chunks = ["et al.","et al.","wt.","etc.","Eq.","Eqs.","eq.","in.","Fig\.","eqs.","i.e.","e.g.","No.","Refs.","Ref.","at.","frac.","vol.","Eqns.","diam.","cf.","et.", "pct."]


def intensified_tokenize(text):
    
    # few angstroms in diameter . 2 Experimental details 2.1 Materi
    conds = re.findall("[^\.]{,10}[a-z]{2,}\.\s[0-9]+\s[A-Z][^\.]{,10}", text)
    for cond in conds:
        rep = re.sub("\.\s","\.\n",cond)
        text = text.replace(cond,rep,1)
        
    conds = re.findall("[^\.]{,10}[A-Z][a-z]{2,}\.\s+[0-9][^\.]{,10}", text)
    for cond in conds:
        rep = re.sub("\.\s+",".",cond)
        text = text.replace(cond,rep,1)
    conds_2 = re.findall("[^\.]{10}[A-Za-z]\.\s[A-Z][^\.]{10}", text)
    for cond in conds_2:
        rep = re.sub("\.\s+",".\n",cond)
        text = text.replace(cond,rep,1)
    sent_tokenize = nltk.sent_tokenize(text)
    all_sents = list()
    for sent in sent_tokenize:
        sub_sents = sent.split("\n")
        if sub_sents:
            for st in sub_sents:
                all_sents.append(st)
        else:
            all_sents.append(sent)

    return all_sents

def sent_process(sent_i,sents,modified_sents,output_sents):
    print(modified_sents)
    print(len(output_sents))
    a_sent = sents[sent_i].strip()
    next_judge = None
    
    if a_sent[-1] != "." and sents[sent_i+1][0].islower():
        new_sent = sents[sent_i] + ' ' + sents[sent_i+1]
        output_sents.append(new_sent)
        del sents[sent_i+1]
        sents[sent_i] = new_sent
        modified_sents += 1
        if sents[sent_i][-1] != "." and sents[sent_i+1][0].islower():
            next_judge = True
    
    elif a_sent[-1] == "." and any(list(a_sent.endswith(chunk) for chunk in error_chunks)):
        new_sent = sents[sent_i] + ' ' +sents[sent_i+1]
        output_sents.append(new_sent)
        del sents[sent_i+1]
        sents[sent_i] = new_sent
        modified_sents += 1
        if sents[sent_i][-1] == "." and any(list(a_sent.endswith(chunk) for chunk in error_chunks)):
            next_judge = True
    else:
        if output_sents and output_sents[-1] != a_sent.strip():
            output_sents.append(a_sent.strip())
        if not output_sents:
            output_sents.append(a_sent.strip())

    return next_judge,sents,modified_sents,output_sents

def sents_reorg(sents):
    """当一句话的结尾不是句号并且下一句话开头不是字母大写，就将这两句话进行拼接"""
    """当一句话的结尾是句号，且结尾是error_chunks中的词，则将两句话进行拼接"""
    
    ori_len = len(sents)
    output_len = 0
    modified_sents = 0
    output_sents = list()
    try:
        for sent_i in range(len(sents)):
            # if output_len + modified_sents != ori_len:
                try:
                    for num in range(5):
                        next_judge,sents,modified_sents,output_sents = sent_process(sent_i,sents,modified_sents,output_sents)
                        if not next_judge:
                            break
                        else:
                            del output_sents[-1]
                            
    
                    # a_sent = sents[sent_i].strip()
                    # if a_sent[-1] != "." and sents[sent_i+1][0].islower():
                    #     new_sent = sents[sent_i] + sents[sent_i+1]
                    #     output_sents.append(new_sent)
                    #     del sents[sent_i+1]
                    # elif a_sent[-1] == "." and any(list(a_sent.endswith(chunk) for chunk in error_chunks)):
                    #     new_sent = sents[sent_i] + sents[sent_i+1]
                        
                    # else:
                    #     output_sents.append(a_sent.strip())
                except Exception as e:
                    print(e)
                    print(sents[sent_i])
                output_len = len(output_sents)
    except Exception as e:
        print(e)
        # else:
            # break
 
    return output_sents


xls = openpyxl.Workbook()
sht_1 = xls.create_sheet(index=0)

# 将所有语料进行分句重组
# with open(r"E:\fasttext\all_exps.txt",mode="r",encoding=("utf-8")) as f:
#     content = f.read()
#     sent_tokenize = intensified_tokenize(content)
#     sents = sents_reorg(sent_tokenize)
#     col_i = 1
#     for sent in sents:
#         sht_1.cell(col_i,1,str(sent))
#         col_i += 1
# xls.save(r'E:\fasttext\reorg_exp_sents.xlsx')

import xlrd
import json

# 将所有语料进行分句重组
xls = xlrd.open_workbook(r"D:\Git\Action_extractor\Action_extractor-main\running_files\positive_paragraphs.xlsx")
sht = xls.sheet_by_index(0)
parags = sht.col_values(0)
for parag in parags:
    sent_tokenize = intensified_tokenize(parag)
    sents = sents_reorg(sent_tokenize)
    col_i = 1
    for sent in sents:
        sht_1.cell(col_i,1,str(sent))
        col_i += 1
xls.save(r'D:\Git\Action_extractor\Action_extractor-main\running_files\reorg_exp_sents.xlsx')


xls = xlrd.open_workbook(r"E:\fasttext\exp_info.xlsx")
sht = xls.sheet_by_index(0)
dois = sht.col_values(0)
parags = sht.col_values(1)
to_json = dict()
for sent_i in range(len(dois)):
    doi = dois[sent_i]
    parag = parags[sent_i]
    sent_tokenize = intensified_tokenize(parag)
    sents = sents_reorg(sent_tokenize)
    if doi not in to_json.keys():
        to_json[doi] = list()
        to_json[doi].extend(sents)
    else:
        to_json[doi].extend(sents)
with open(r"E:\fasttext\add_exp_model\parag_info.json","w",encoding="utf-8") as file:
    file.write(json.dumps(to_json))

