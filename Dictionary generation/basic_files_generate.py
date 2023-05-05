
import os
from nltk.parse.stanford import StanfordParser
from tqdm import tqdm
from chemdataextractor.doc import Paragraph
from dictionary import Dictionary
from sent_token_modified import intensified_tokenize,sents_reorg
import xlrd
import json
import openpyxl

def sent_constituent_parsing(sent,java_path,stanford_parser_path,stanford_model_path):
    parsing_result = None
    if 40<=len(sent)<=1000:
        os.environ["JAVAHOME"] = java_path
        scp = StanfordParser(path_to_jar=(stanford_parser_path),
                             path_to_models_jar=(stanford_model_path))
        result = list(scp.raw_parse(sent))
        parsing_result = result[0]
    return parsing_result


# Reorganize all the corpus into clauses, lowercase processing, and syntactic parsing
class Basic_file:
    """
    parags_path: xlsx file contain positive paragraph in first column, article DOI in second column;
    reorg_save_path: xlsx file contian parags in first column;
    lower_save_path: txt file contain each sentence with initial lowercase processing;
    java_path: the path of java.exe
    stanford_parser_path: the path of stanford-parser.jar;
    stanford_model_path: the path of stanford-parser-x.x.x-models.jar;
    parsing_save_path: the path of parsing results of every sentence;
    c_path: the path of configuration file;

    """
    def __init__(self,parags_path, reorg_save_path, lower_save_path, java_path, stanford_parser_path,
                 stanford_model_path, parsing_save_path, c_path):

        self.parags_path = parags_path
        self.reorg_save_path = reorg_save_path
        self.lower_save_path = lower_save_path
        self.java_path = java_path
        self.stanford_parser_path = stanford_parser_path
        self.stanford_model_path = stanford_model_path
        self.parsing_save_path = parsing_save_path
        self.c_path = c_path


    def reorg_lower(self):
        error_chunks = Dictionary(self.c_path).error_chunks
        xls_n = openpyxl.Workbook()
        sht_1 = xls_n.create_sheet(index=0)
        xls = xlrd.open_workbook(self.parags_path)
        sht = xls.sheet_by_index(0)
        parags = sht.col_values(0)
        lower_sents = ""
        reorg_sents = list()
        print("**Start process paragraphs.**")
        for parag in tqdm(parags):
            sent_tokenize = intensified_tokenize(parag)
            sents = sents_reorg(sent_tokenize,error_chunks)
            col_i = 1
            for sent in sents:
                sht_1.cell(col_i, 1, str(sent))
                reorg_sents.append(str(sent))
                col_i += 1
                lower_sents += sent
                lower_sents += "\n"
        xls_n.save(self.reorg_save_path)
        with open(self.lower_save_path, "w+", encoding="utf-8") as l_f:
            l_f.write(lower_sents)
        return reorg_sents

    def constituent_parsing(self,reorg_sents):

        error_sents = list()
        parsing_dict = dict()
        print("**Start constituent parsing.**")
        for parag in tqdm(reorg_sents):
            para = Paragraph(parag)
            sentences = para.sentences
            sents = list()
            for sentence in sentences:
                sents.append(sentence.text)
            for sent in sents:
                try:
                    sent_p = sent_constituent_parsing(sent, self.java_path, self.stanford_parser_path,
                                                      self.stanford_model_path)
                    parsing_dict[sent] = str(sent_p)
                except Exception as e:
                    print(e)
                    print(sent)
                    error_sents.append(sent)
        with open(self.parsing_save_path, "w+",encoding="utf-8") as file:
            json.dump(parsing_dict, file)
        return error_sents



parags_path = r".\positive_paragraphs.xlsx"
reorg_save_path = r'.\reorg_sents.xlsx'
lower_save_path = r'.\lower_corpus.txt'
java_path = r"./jre1.8.0_321/bin/java.exe"
stanford_parser_path = r"./stanford-parser-full-2020-11-17/stanford-parser.jar"
stanford_model_path = r"./stanford-parser-full-2020-11-17/stanford-parser-4.2.0-models.jar"
parsing_save_path = r".\parsing_results.json"
c_path = r"dictionary.ini"

bf = Basic_file(parags_path, reorg_save_path, lower_save_path, java_path, stanford_parser_path,
                 stanford_model_path, parsing_save_path, c_path)
reorg_sents = bf.reorg_lower()
error_sents = bf.constituent_parsing(reorg_sents)
