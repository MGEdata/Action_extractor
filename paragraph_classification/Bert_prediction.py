# -*- coding: utf-8 -*-
"""
Created on Thu Aug  5 20:43:14 2021

@author: win
"""
# -*- coding: utf-8 -*-
"""
Created on Thu Aug  5 18:28:26 2021

@author: win
"""
from sklearn.feature_extraction.text import TfidfVectorizer
import joblib
from transformers import logging
logging.set_verbosity_warning()
from sklearn.model_selection import train_test_split
import random
import torch
from torch.utils.data import TensorDataset, DataLoader, random_split
from transformers import BertTokenizer, BertModel
from transformers import BertForSequenceClassification, AdamW
from transformers import get_linear_schedule_with_warmup
import numpy as np
import pandas as pd
import os
from sklearn.metrics import accuracy_score, f1_score, recall_score
import re
import nltk
from gensim.models.doc2vec import TaggedDocument
import multiprocessing
from sklearn.linear_model import LogisticRegression


def tokenize_text(text):
    tokens = []
    for sent in nltk.sent_tokenize(text):
        for word in nltk.word_tokenize(sent):
            tokens.append(word)
    return tokens


def vec_for_learning(model, tagged_docs):
    sents = tagged_docs.values
    targets, regressors = zip(*[(doc.tags[0], model.infer_vector(doc.words, steps=20)) for doc in sents])
    return targets, regressors




def ber_embed(model_path, doc_list):
    text = doc_list[0]
    with torch.no_grad():
        tokenizer = BertTokenizer.from_pretrained(model_path)
        model = BertModel.from_pretrained(model_path)
        # model = model.cuda()
        # model = nn.DataParallel(model,device_ids=range(torch.cuda.device_count()))
        input_ids = tokenizer(text,truncation=True, padding=True, max_length=512, return_tensors='pt')
        outputs = model(**input_ids)
        pooled_output = outputs[1]
    return pooled_output

def get_vectors(bert_path, tagged_docs, doc_label):
    sents = tagged_docs.values
    outcome = list()
    # infer_vec = model.infer_vector(doc.words, steps=5)
    for doc in sents:
        doc_list = list()
        doc_text = " ".join(doc.words)
        doc_list.append(doc_text)
        model_infer = ber_embed(bert_path, doc_list)
        model_infer = model_infer.detach().numpy()[0]

        if "experiment" in doc_label[doc_text].lower() or "method" in doc_label[doc_text].lower():
            tag_vec = np.ones(1)
            model_i = np.concatenate((model_infer, tag_vec), axis=0)
        else:
            tag_vec = np.zeros(1)
            model_i = np.concatenate((model_infer, tag_vec), axis=0)
            # print(model_i.shape)
        outcome.append((doc.tags[0], model_i))
    targets, regressors = zip(*outcome)
    # targets, regressors = zip(*[(doc.tags[0], model.infer_vector(doc.words, steps=20)) for doc in sents])
    return targets, regressors


def new_model_train(bert_path, corpus_path, save_path):
    """
    在确定好模型类别之后，将所有的标记语料用于训练模型，保存模型。
    :return:
    """
    csv_data = pd.read_csv(corpus_path)
    df = pd.DataFrame(csv_data)
    df_shape = df.shape
    length = df_shape[0]
    outcome = {}
    for i in range(length):
        example_i = df.iloc[i]
        if len(example_i) > 0:
            outcome.update({example_i[0]: example_i[1]})
    paragraph_list = list(outcome.keys())
    english_punctuations = [',', '.', ':', ';', '?', '(', ')', '[', ']', '&', '!', '*', '@', '#', '$', '%']
    words_list = []
    outcome_1 = {}
    all_corpus = list()
    doc_label = dict()  # 用于记录每一个自然段对应的标签
    for i in range(len(paragraph_list)):
        r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        paragraph_0 = re.sub(r, ' ', paragraph_list[i].lower())
        words_list_0 = nltk.word_tokenize(paragraph_0)
        words_list.append(words_list_0)
        keys = " ".join(words_list[i])
        all_corpus.append(keys)
        outcome_1.update({keys: df.iloc[i][1]})
        doc_label.update({keys: df.iloc[i][2]})
    # 可以去重
    df = pd.DataFrame(pd.Series(outcome_1), columns=['y'])  # 将字典转化为dataframe的形式,第一列会变成索引列
    train = df.reset_index().rename(columns={'index': 'x'})
    train_tagged = train.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)

    cores = multiprocessing.cpu_count()
    y_train, X_train = get_vectors(bert_path, train_tagged, doc_label)
    log = LogisticRegression(max_iter=1000, class_weight='balanced').fit(y_train, X_train)
    joblib.dump(log, save_path)


def get_accuracy_f1(lable_path, bert_path, rs):
    csv_data = pd.read_csv(lable_path)
    df = pd.DataFrame(csv_data)
    df_shape = df.shape
    length = df_shape[0]
    outcome = {}
    for i in range(length):
        example_i = df.iloc[i]
        if len(example_i) > 0:
            outcome.update({example_i[0]: example_i[1]})
    paragraph_list = list(outcome.keys())
    english_punctuations = [',', '.', ':', ';', '?', '(', ')', '[', ']', '&', '!', '*', '@', '#', '$', '%']
    words_list = []
    outcome_1 = {}
    all_corpus = list()
    doc_label = dict()  # 用于记录每一个自然段对应的标签
    for i in range(len(paragraph_list)):
        r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        paragraph_0 = re.sub(r, ' ', paragraph_list[i])
        words_list_0 = nltk.word_tokenize(paragraph_0)
        words_list.append(words_list_0)
        keys = " ".join(words_list[i])
        all_corpus.append(keys)
        outcome_1.update({keys: df.iloc[i][1]})
        doc_label.update({keys: df.iloc[i][2]})
    # 可以去重
    df = pd.DataFrame(pd.Series(outcome_1), columns=['y'])  # 将字典转化为dataframe的形式,第一列会变成索引列
    df = df.reset_index().rename(columns={'index': 'x'})

    train, test = train_test_split(df, test_size=0.2, random_state=rs)  # 分为训练集和测试集

    train_tagged = train.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)
    test_tagged = test.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)
    cores = multiprocessing.cpu_count()

    y_train, X_train = get_vectors(bert_path, train_tagged, doc_label)
    y_test, X_test = get_vectors(bert_path, test_tagged, doc_label)

    log = LogisticRegression(max_iter=1000, class_weight='balanced').fit(X_train, y_train)
    predictions = log.predict_proba(X_test)
    y_pred = log.predict(X_test)
    accuracy = accuracy_score(y_test, y_pred)
    score = f1_score(y_test, y_pred, average='weighted')
    # print('Testing accuracy %s' % accuracy_score(y_test, y_pred))
    # print('Testing F1 score: {}'.format(f1_score(y_test, y_pred, average='weighted')))
    return accuracy, score


def predict_on_file(bert_path, label_path, target_excel, predict_results_path, model_save_path):
    """
    训练段落分类模型并在新语料中预测，结果输出到新的excel文件中
    :param bert_path: 训练好的模型被保存到的地址
    :param label_path: 标记的数据所储存的地址
    :param target_excel: 要预测的新语料所在地址
    :param predict_results_path: 预测结果的输出地址
    :return:
    """
    csv_data = pd.read_csv(label_path)
    df = pd.DataFrame(csv_data)
    df_shape = df.shape
    length = df_shape[0]
    outcome = {}
    for i in range(length):
        example_i = df.iloc[i]
        if len(example_i) > 0:
            outcome.update({example_i[0]: example_i[1]})
    paragraph_list = list(outcome.keys())
    english_punctuations = [',', '.', ':', ';', '?', '(', ')', '[', ']', '&', '!', '*', '@', '#', '$', '%']
    words_list = []
    outcome_1 = {}
    all_corpus = list()
    doc_label = dict()  # 用于记录每一个自然段对应的标签
    for i in range(len(paragraph_list)):
        r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        paragraph_0 = re.sub(r, ' ', paragraph_list[i].lower())
        words_list_0 = nltk.word_tokenize(paragraph_0)
        words_list.append(words_list_0)
        keys = " ".join(words_list[i])
        all_corpus.append(keys)
        outcome_1.update({keys: df.iloc[i][1]})
        doc_label.update({keys: df.iloc[i][2]})
    # 可以去重
    df = pd.DataFrame(pd.Series(outcome_1), columns=['y'])  # 将字典转化为dataframe的形式,第一列会变成索引列
    df = df.reset_index().rename(columns={'index': 'x'})
    train = df
    train_tagged = train.apply(
        lambda r: TaggedDocument(words=tokenize_text(r['x']), tags=[r.y]), axis=1)
    cores = multiprocessing.cpu_count()
    y_test, X_test = get_vectors(bert_path, train_tagged, doc_label)
    log = LogisticRegression(max_iter=1000, class_weight='balanced').fit(X_test, y_test)
    joblib.dump(log, model_save_path) 

    # 利用本地向量进行预测
    import xlrd
    new_data = xlrd.open_workbook(target_excel) 
    sht = new_data.sheet_by_index(1)
    rows = sht.nrows
    col_0 = sht.col_values(0)
    col_1 = sht.col_values(1)
    outcome = dict()
    outcome_v = list()
    doi_outcome = dict()
    for i in range(rows):
        words_list = []
        example_i = sht.row_values(i)
        if len(example_i) > 0:
            outcome.update({example_i[1]: example_i[0]})
            text = example_i[1]
            r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
            paragraph_ = re.sub(r, ' ', text.lower())
            words_list_0 = nltk.word_tokenize(paragraph_)
            keys = " ".join(words_list_0)
            doi_outcome.update({keys: example_i[2]})
    paragraph_list = list(outcome.keys())
    words_list = list()
    doc_label = dict()

    for i in range(len(paragraph_list)):
        r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        paragraph_0 = re.sub(r, ' ', paragraph_list[i].lower())
        words_list_0 = nltk.word_tokenize(paragraph_0)
        words_list.append(words_list_0)
        keys = " ".join(words_list[i])
        doc_label.update({keys: col_1[i]})

    import openpyxl
    xls = openpyxl.Workbook()
    sht = xls.create_sheet(index=0)
    col_i = 1
    for words in words_list:
        doc_text = " ".join(words)
        # model_infer = new_model.infer_vector(words, steps=20)

        doc_list = list()
        doc_list.append(doc_text)
        model_infer = ber_embed(model_save_path, doc_list)
        model_infer = model_infer.detach().numpy()[0]

        if "experiment" in doc_label[doc_text].lower() or "method" in doc_label[doc_text].lower():
            tag_vec = np.ones(1)
            model_i = np.concatenate((model_infer, tag_vec), axis=0)
        else:
            tag_vec = np.zeros(1)
            model_i = np.concatenate((model_infer, tag_vec), axis=0)
        outcome_v.append(model_i)
        sht.cell(col_i, 1, paragraph_list[col_i - 1])
        doi = doi_outcome[doc_text]
        sht.cell(col_i, 3, doi)
        col_i += 1
    predict_outcome = log.predict(tuple(outcome_v))
    pre_prob = log.predict_proba(tuple(outcome_v)).tolist()
    col_i = 1
    for tag in predict_outcome:
        sht.cell(col_i, 2, tag)  # 写入表格的只能是字符型数据
        sht.cell(col_i, 4, str(pre_prob[col_i - 1]))
        col_i += 1
    xls.save(predict_results_path)  
    # log_model = joblib.load(model_path)


predict_on_file(r"./model/bert_model",
                r"./label.csv",
                r"./all_paragraphs_predicted.xlsx",
                r"./bert_predict_results.xlsx",
                r"./model/bert_logist")

